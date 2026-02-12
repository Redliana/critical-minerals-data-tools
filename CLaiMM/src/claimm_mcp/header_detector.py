"""
CSV/Excel Header Detection Tool

Detects column headers from CSV and Excel files without downloading the full file.
Uses HTTP Range requests to fetch only the first portion of the file.
"""

from __future__ import annotations

import asyncio
import csv
import io
from typing import Any

import httpx

from .config import get_settings


class HeaderDetector:
    """Detect headers from tabular files without full download."""

    def __init__(self):
        self.settings = get_settings()
        self.headers = {
            "X-CKAN-API-Key": self.settings.edx_api_key,
            "User-Agent": "EDX-USER",
        }
        # Bytes to fetch for header detection
        self.csv_fetch_bytes = 8192  # 8KB should be enough for headers
        self.xlsx_fetch_bytes = 65536  # 64KB for Excel (need more for structure)

    async def detect_csv_headers(
        self,
        resource_id: str,
        delimiter: str | None = None,
    ) -> dict[str, Any]:
        """
        Detect CSV column headers by fetching only the first portion of the file.

        Args:
            resource_id: EDX resource ID
            delimiter: CSV delimiter (auto-detected if not provided)

        Returns:
            Dict with headers, sample_rows, detected_types, and metadata
        """
        download_url = f"https://edx.netl.doe.gov/resource/{resource_id}/download"

        async with httpx.AsyncClient(timeout=30.0) as client:
            # First, try a Range request
            range_headers = {
                **self.headers,
                "Range": f"bytes=0-{self.csv_fetch_bytes - 1}",
            }

            try:
                response = await client.get(
                    download_url,
                    headers=range_headers,
                    follow_redirects=True,
                )

                # Check if server supports range requests
                if response.status_code == 206:  # Partial Content
                    content = response.text
                    partial = True
                elif response.status_code == 200:
                    # Server doesn't support range, but returned full content
                    # Only use first portion
                    content = response.text[: self.csv_fetch_bytes]
                    partial = len(response.text) > self.csv_fetch_bytes
                else:
                    return {
                        "success": False,
                        "error": f"HTTP {response.status_code}",
                        "resource_id": resource_id,
                    }

            except Exception as e:
                return {
                    "success": False,
                    "error": str(e),
                    "resource_id": resource_id,
                }

        # Parse CSV content
        return self._parse_csv_content(content, resource_id, delimiter, partial)

    def _parse_csv_content(
        self,
        content: str,
        resource_id: str,
        delimiter: str | None,
        partial: bool,
    ) -> dict[str, Any]:
        """Parse CSV content and extract headers and sample data."""
        lines = content.strip().split("\n")

        if not lines:
            return {
                "success": False,
                "error": "Empty content",
                "resource_id": resource_id,
            }

        # Auto-detect delimiter if not provided
        if delimiter is None:
            first_line = lines[0]
            delimiter = self._detect_delimiter(first_line)

        # Parse with csv module
        try:
            reader = csv.reader(io.StringIO(content), delimiter=delimiter)
            rows = list(reader)
        except Exception as e:
            return {
                "success": False,
                "error": f"CSV parse error: {e}",
                "resource_id": resource_id,
            }

        if not rows:
            return {
                "success": False,
                "error": "No rows parsed",
                "resource_id": resource_id,
            }

        headers = rows[0]
        sample_rows = rows[1:6] if len(rows) > 1 else []  # Up to 5 sample rows

        # Detect column types from sample data
        column_types = self._detect_column_types(headers, sample_rows)

        return {
            "success": True,
            "resource_id": resource_id,
            "partial_download": partial,
            "bytes_fetched": len(content.encode("utf-8")),
            "delimiter": delimiter,
            "column_count": len(headers),
            "headers": headers,
            "column_types": column_types,
            "sample_rows": sample_rows,
            "rows_sampled": len(sample_rows),
        }

    def _detect_delimiter(self, line: str) -> str:
        """Auto-detect CSV delimiter."""
        delimiters = [",", "\t", ";", "|"]
        counts = {d: line.count(d) for d in delimiters}
        return max(counts, key=counts.get) if max(counts.values()) > 0 else ","

    def _detect_column_types(
        self,
        headers: list[str],
        sample_rows: list[list[str]],
    ) -> list[dict[str, Any]]:
        """Infer column types from sample data."""
        if not sample_rows:
            return [{"name": h, "type": "unknown"} for h in headers]

        column_types = []
        for i, header in enumerate(headers):
            values = [row[i] if i < len(row) else "" for row in sample_rows]
            values = [v for v in values if v.strip()]  # Non-empty values

            col_type = self._infer_type(values)
            column_types.append(
                {
                    "name": header,
                    "type": col_type["type"],
                    "nullable": any(
                        not v.strip()
                        for v in [row[i] if i < len(row) else "" for row in sample_rows]
                    ),
                    "sample_values": values[:3],
                    **col_type.get("metadata", {}),
                }
            )

        return column_types

    def _infer_type(self, values: list[str]) -> dict[str, Any]:
        """Infer data type from sample values."""
        if not values:
            return {"type": "unknown"}

        # Check for numeric
        numeric_count = 0
        int_count = 0
        float_count = 0

        for v in values:
            v = v.strip().replace(",", "")  # Handle thousands separator
            try:
                float(v)
                numeric_count += 1
                if "." in v or "e" in v.lower():
                    float_count += 1
                else:
                    int_count += 1
            except ValueError:
                pass

        if numeric_count == len(values):
            if float_count > 0:
                return {"type": "float", "metadata": {"precision": "double"}}
            return {"type": "integer"}

        # Check for date/datetime patterns
        date_patterns = ["-", "/"]
        if all(any(p in v for p in date_patterns) for v in values):
            if any(":" in v for v in values):
                return {"type": "datetime"}
            return {"type": "date"}

        # Check for boolean
        bool_values = {"true", "false", "yes", "no", "1", "0", "t", "f", "y", "n"}
        if all(v.lower().strip() in bool_values for v in values):
            return {"type": "boolean"}

        # Default to string
        max_len = max(len(v) for v in values) if values else 0
        return {"type": "string", "metadata": {"max_length": max_len}}

    async def detect_xlsx_headers(self, resource_id: str) -> dict[str, Any]:
        """
        Detect Excel headers. Note: Excel files require more data due to format.
        For large files, this may need to download more than just headers.

        Args:
            resource_id: EDX resource ID

        Returns:
            Dict with headers and sheet information
        """
        download_url = f"https://edx.netl.doe.gov/resource/{resource_id}/download"

        async with httpx.AsyncClient(timeout=60.0) as client:
            # Excel files need more data due to their structure
            # Try Range request first
            range_headers = {
                **self.headers,
                "Range": f"bytes=0-{self.xlsx_fetch_bytes - 1}",
            }

            try:
                response = await client.get(
                    download_url,
                    headers=range_headers,
                    follow_redirects=True,
                )

                if response.status_code in [200, 206]:
                    content = response.content
                else:
                    return {
                        "success": False,
                        "error": f"HTTP {response.status_code}",
                        "resource_id": resource_id,
                    }

            except Exception as e:
                return {
                    "success": False,
                    "error": str(e),
                    "resource_id": resource_id,
                }

        # Try to parse Excel - this may fail for partial downloads
        try:
            import openpyxl

            # For partial downloads, Excel parsing often fails
            # We need the full file for reliable parsing
            wb = openpyxl.load_workbook(
                io.BytesIO(content),
                read_only=True,
                data_only=True,
            )

            sheets = {}
            for sheet_name in wb.sheetnames:
                ws = wb[sheet_name]
                rows = list(ws.iter_rows(max_row=6, values_only=True))
                if rows:
                    headers = [str(c) if c else "" for c in rows[0]]
                    sample_rows = [[str(c) if c else "" for c in row] for row in rows[1:6]]
                    sheets[sheet_name] = {
                        "headers": headers,
                        "column_count": len(headers),
                        "sample_rows": sample_rows,
                    }

            wb.close()

            return {
                "success": True,
                "resource_id": resource_id,
                "format": "XLSX",
                "bytes_fetched": len(content),
                "sheets": sheets,
                "sheet_names": list(sheets.keys()),
            }

        except Exception as e:
            return {
                "success": False,
                "error": f"Excel parse error (may need full file): {e}",
                "resource_id": resource_id,
                "bytes_fetched": len(content),
                "suggestion": "Excel files often require full download for reliable parsing",
            }

    async def detect_headers(
        self,
        resource_id: str,
        format: str | None = None,
    ) -> dict[str, Any]:
        """
        Auto-detect headers based on file format.

        Args:
            resource_id: EDX resource ID
            format: File format (CSV, XLSX, etc.) - auto-detected if not provided

        Returns:
            Dict with detected headers and metadata
        """
        format = (format or "").upper()

        if format == "CSV" or not format:
            result = await self.detect_csv_headers(resource_id)
            if result["success"] or format == "CSV":
                return result

        if format in ["XLSX", "XLS", "XLSM"] or not format:
            return await self.detect_xlsx_headers(resource_id)

        return {
            "success": False,
            "error": f"Unsupported format: {format}",
            "resource_id": resource_id,
        }


async def detect_all_csv_headers(resource_ids: list[str]) -> list[dict[str, Any]]:
    """
    Detect headers for multiple CSV resources in parallel.

    Args:
        resource_ids: List of EDX resource IDs

    Returns:
        List of detection results
    """
    detector = HeaderDetector()
    tasks = [detector.detect_csv_headers(rid) for rid in resource_ids]
    return await asyncio.gather(*tasks)
